from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from datetime import datetime
from io import BytesIO

def create_styled_workbook(title):
    """Create a new workbook with basic styling"""
    wb = Workbook()
    ws = wb.active
    ws.title = title
    return wb, ws

def style_header_row(ws, row_num, columns):
    """Apply styling to header row"""
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    alignment = Alignment(horizontal="center", vertical="center")
    
    for col_num, column_name in enumerate(columns, 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = column_name
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = alignment

def add_school_header(ws, title, subtitle=None):
    """Add school branding header to worksheet"""
    ws.merge_cells('A1:F1')
    ws['A1'] = 'EECOHM School of Excellence - Canteen Management'
    ws['A1'].font = Font(bold=True, size=14, color="1E293B")
    ws['A1'].alignment = Alignment(horizontal="center")
    
    ws.merge_cells('A2:F2')
    ws['A2'] = title
    ws['A2'].font = Font(bold=True, size=12, color="4F46E5")
    ws['A2'].alignment = Alignment(horizontal="center")
    
    if subtitle:
        ws.merge_cells('A3:F3')
        ws['A3'] = subtitle
        ws['A3'].font = Font(size=10, color="64748B")
        ws['A3'].alignment = Alignment(horizontal="center")
        return 5  # Data starts at row 5
    
    return 4  # Data starts at row 4

def auto_size_columns(ws):
    """Auto-size columns based on content"""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

def generate_transactions_excel(transactions, filters=None):
    """Generate Excel file for transactions list"""
    wb, ws = create_styled_workbook("Transactions")
    
    # Add header
    filter_text = ""
    if filters:
        parts = []
        if filters.get('start_date'):
            parts.append(f"From: {filters['start_date']}")
        if filters.get('end_date'):
            parts.append(f"To: {filters['end_date']}")
        if filters.get('payment_type'):
            parts.append(f"Payment: {filters['payment_type']}")
        filter_text = " | ".join(parts) if parts else "All Transactions"
    
    start_row = add_school_header(ws, "Transaction Report", filter_text)
    
    # Headers
    columns = ['ID', 'Date', 'Time', 'Cashier', 'Payment Type', 'Items', 'Total Amount', 'Status']
    style_header_row(ws, start_row, columns)
    
    # Data
    row_num = start_row + 1
    for tx in transactions:
        ws.cell(row=row_num, column=1, value=tx.id)
        ws.cell(row=row_num, column=2, value=tx.timestamp.strftime('%Y-%m-%d'))
        ws.cell(row=row_num, column=3, value=tx.timestamp.strftime('%H:%M:%S'))
        ws.cell(row=row_num, column=4, value=tx.cashier.full_name)
        ws.cell(row=row_num, column=5, value=tx.payment_type.upper())
        ws.cell(row=row_num, column=6, value=tx.lines.count())
        ws.cell(row=row_num, column=7, value=float(tx.total_amount))
        ws.cell(row=row_num, column=8, value='CANCELED' if tx.is_canceled else 'ACTIVE')
        row_num += 1
    
    # Format currency column
    for row in range(start_row + 1, row_num):
        ws.cell(row=row, column=7).number_format = 'Rs. #,##0.00'
    
    auto_size_columns(ws)
    
    # Return as HTTP response
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'transactions_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response

def generate_daily_summary_excel(date, summary_data):
    """Generate Excel file for daily summary report"""
    wb, ws = create_styled_workbook("Daily Summary")
    
    # Add header
    start_row = add_school_header(ws, "Daily Summary Report", f"Date: {date}")
    
    # Summary section
    ws.cell(row=start_row, column=1, value="Metric").font = Font(bold=True)
    ws.cell(row=start_row, column=2, value="Value").font = Font(bold=True)
    
    metrics = [
        ("Total Sales", f"Rs. {summary_data.get('total_sales', 0):.2f}"),
        ("Total Transactions", summary_data.get('transaction_count', 0)),
        ("Cash Sales", f"Rs. {summary_data.get('cash_sales', 0):.2f}"),
        ("Credit Sales", f"Rs. {summary_data.get('credit_sales', 0):.2f}"),
        ("Average Transaction", f"Rs. {summary_data.get('avg_transaction', 0):.2f}"),
    ]
    
    for idx, (metric, value) in enumerate(metrics, start_row + 1):
        ws.cell(row=idx, column=1, value=metric)
        ws.cell(row=idx, column=2, value=value)
    
    auto_size_columns(ws)
    
    # Return as HTTP response
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'daily_summary_{date}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response

def generate_monthly_summary_excel(year, month, summary_data):
    """Generate Excel file for monthly summary report"""
    wb, ws = create_styled_workbook("Monthly Summary")
    
    # Add header
    month_name = datetime(year, month, 1).strftime('%B %Y')
    start_row = add_school_header(ws, "Monthly Summary Report", month_name)
    
    # Summary section
    ws.cell(row=start_row, column=1, value="Metric").font = Font(bold=True)
    ws.cell(row=start_row, column=2, value="Value").font = Font(bold=True)
    
    metrics = [
        ("Total Sales", f"Rs. {summary_data.get('total_sales', 0):.2f}"),
        ("Total Transactions", summary_data.get('transaction_count', 0)),
        ("Cash Sales", f"Rs. {summary_data.get('cash_sales', 0):.2f}"),
        ("Credit Sales", f"Rs. {summary_data.get('credit_sales', 0):.2f}"),
        ("Total Expenses", f"Rs. {summary_data.get('total_expenses', 0):.2f}"),
        ("Net Profit", f"Rs. {summary_data.get('net_profit', 0):.2f}"),
    ]
    
    for idx, (metric, value) in enumerate(metrics, start_row + 1):
        ws.cell(row=idx, column=1, value=metric)
        ws.cell(row=idx, column=2, value=value)
    
    auto_size_columns(ws)
    
    # Return as HTTP response
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'monthly_summary_{year}_{month:02d}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response
